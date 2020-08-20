from openpyxl import load_workbook

file_path = ''
sheet_name = ''


# Get Param Value by Param Name
def get_param_value(param_name):
    global file_path, sheet_name

    # Load in the workbook
    wb = load_workbook(file_path)

    # Get a sheet by name
    sheet = wb[sheet_name]

    for row in range(sheet.min_row, sheet.max_row + 1):
        if sheet.cell(row, 1).value.upper() == param_name.upper():
            cell_value = sheet.cell(row, 2).value
            return cell_value
            break
